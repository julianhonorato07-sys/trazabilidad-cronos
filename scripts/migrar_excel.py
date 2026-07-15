#!/usr/bin/env python3
"""Migra el Excel "Trazabilidad de Cronos .xlsx" a los datos de la app.

Genera:
  - src/data/seed.json   -> datos de arranque del modo demo (localStorage)
  - supabase/seed.sql    -> inserts para la base Postgres/Supabase (con --sql)

Uso:
  python3 scripts/migrar_excel.py [ruta_al_excel] [--sql]
"""
import datetime
import json
import pathlib
import re
import sys

import openpyxl

RAIZ = pathlib.Path(__file__).resolve().parent.parent
args = [a for a in sys.argv[1:] if not a.startswith("--")]
SRC = args[0] if args else str(pathlib.Path.home() / "Downloads" / "Trazabilidad de Cronos .xlsx")
CON_SQL = "--sql" in sys.argv

TIPOS_FALLA = [
    {"id": 1, "nombre": "Bollo", "usa_particularidad": True, "requiere_particularidad": False},
    {"id": 2, "nombre": "Falta de particularidad", "usa_particularidad": True, "requiere_particularidad": True},
    {"id": 3, "nombre": "Para repintar", "usa_particularidad": False, "requiere_particularidad": False},
    {"id": 4, "nombre": "Ya repintada", "usa_particularidad": False, "requiere_particularidad": False},
    {"id": 5, "nombre": "JyP", "usa_particularidad": True, "requiere_particularidad": False},
    {"id": 6, "nombre": "Marca de operación", "usa_particularidad": True, "requiere_particularidad": False},
    {"id": 7, "nombre": "Caso especial", "usa_particularidad": False, "requiere_particularidad": False},
    {"id": 8, "nombre": "Otra anomalía", "usa_particularidad": False, "requiere_particularidad": False},
]
# Particularidades (partes) por tipo de unidad. Cronos son las históricas del Excel (ids 1-9,
# no tocar para no romper la migración). Cabina y Caja (KP1) son TENTATIVAS: confirmar con planta.
PARTICULARIDADES = [
    {"id": 1, "codigo": "CAPOT", "nombre": "Capot", "tipo": "cronos"},
    {"id": 2, "codigo": "BAUL", "nombre": "Baúl", "tipo": "cronos"},
    {"id": 3, "codigo": "PASX", "nombre": "Puerta anterior izq.", "tipo": "cronos"},
    {"id": 4, "codigo": "PADX", "nombre": "Puerta anterior der.", "tipo": "cronos"},
    {"id": 5, "codigo": "PPSX", "nombre": "Puerta posterior izq.", "tipo": "cronos"},
    {"id": 6, "codigo": "PPDX", "nombre": "Puerta posterior der.", "tipo": "cronos"},
    {"id": 7, "codigo": "GBSX", "nombre": "Guardabarro izq.", "tipo": "cronos"},
    {"id": 8, "codigo": "GBDX", "nombre": "Guardabarro der.", "tipo": "cronos"},
    {"id": 9, "codigo": "ZOC", "nombre": "Zócalo", "tipo": "cronos"},
    {"id": 10, "codigo": "TECHO", "nombre": "Techo", "tipo": "cronos"},
    # Cabina (KP1) — TENTATIVO
    {"id": 11, "codigo": "CAB-PSX", "nombre": "Puerta izq.", "tipo": "cabina"},
    {"id": 12, "codigo": "CAB-PDX", "nombre": "Puerta der.", "tipo": "cabina"},
    {"id": 13, "codigo": "CAB-CAPOT", "nombre": "Capot", "tipo": "cabina"},
    {"id": 14, "codigo": "CAB-TECHO", "nombre": "Techo", "tipo": "cabina"},
    {"id": 15, "codigo": "CAB-PARSX", "nombre": "Parante izq.", "tipo": "cabina"},
    {"id": 16, "codigo": "CAB-PARDX", "nombre": "Parante der.", "tipo": "cabina"},
    {"id": 17, "codigo": "CAB-GBSX", "nombre": "Guardabarro izq.", "tipo": "cabina"},
    {"id": 18, "codigo": "CAB-GBDX", "nombre": "Guardabarro der.", "tipo": "cabina"},
    # Caja (KP1) — TENTATIVO
    {"id": 19, "codigo": "CAJ-LSX", "nombre": "Lateral izq.", "tipo": "caja"},
    {"id": 20, "codigo": "CAJ-LDX", "nombre": "Lateral der.", "tipo": "caja"},
    {"id": 21, "codigo": "CAJ-PORT", "nombre": "Portón trasero", "tipo": "caja"},
    {"id": 22, "codigo": "CAJ-PISO", "nombre": "Piso de caja", "tipo": "caja"},
    {"id": 23, "codigo": "CAJ-FRENTE", "nombre": "Frente de caja", "tipo": "caja"},
    {"id": 24, "codigo": "CAJ-BARSX", "nombre": "Baranda izq.", "tipo": "caja"},
    {"id": 25, "codigo": "CAJ-BARDX", "nombre": "Baranda der.", "tipo": "caja"},
]
OPERARIOS = [
    {"id": 1, "nombre": "Julián H.", "rol": "supervisor", "activo": True},
    {"id": 2, "nombre": "Revisión 1", "rol": "revision", "activo": True},
    {"id": 3, "nombre": "Revisión 2", "rol": "revision", "activo": True},
    {"id": 4, "nombre": "OLEO 1", "rol": "oleo", "activo": True},
    {"id": 5, "nombre": "OLEO 2", "rol": "oleo", "activo": True},
    {"id": 6, "nombre": "Box 1", "rol": "box", "activo": True},
    {"id": 7, "nombre": "Box 2", "rol": "box", "activo": True},
]
# Colores confirmados en planta (12/07/2026). El hex es el swatch visual de la app.
CEST_INFO = {
    "249": ("Blanco Banchisa", "#F0EEE6"),
    "534": ("Blanco Alaska", "#FBFBFA"),
    "619": ("Plata Bari", "#C9CBCD"),
    "806": ("Negro", "#1E2124"),
    "978": ("Rojo", "#9E2B32"),
    "979": ("Gris Silverstone", "#6A6E72"),
}
MAPA_PART = {"CAPOT": 1, "BAUL": 2, "PASX": 3, "PADX": 4, "PPSX": 5, "PPDX": 6}


def iso(v):
    return v.isoformat() if isinstance(v, datetime.datetime) else None


def norm_part(v):
    s = str(v).strip().upper().replace("Ú", "U").replace("Á", "A")
    return MAPA_PART.get(s)


wb = openpyxl.load_workbook(SRC, data_only=True)

# 1 · Maestro de unidades desde el export del sistema de planta
unidades, por_cis = [], {}
for r in wb["CISDetailsDataActual"].iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    cis = str(r[0]).strip()
    u = {
        "id": len(unidades) + 1, "cis": cis, "check_digit": r[1],
        "salida_linea": iso(r[2]), "st": str(r[3] or "").strip(),
        "ssc": str(r[4] or ""), "smo": str(r[5] or ""), "cest": str(r[6] or ""),
        "tipo": "cronos",
    }
    unidades.append(u)
    por_cis[cis] = u

# 2 · Secciones de la hoja Análisis (categorías numeradas "N · TÍTULO")
ana = wb["Análisis"]
rows = [[c.value for c in fila] for fila in ana.iter_rows()]
secciones = {}
for i, r in enumerate(rows):
    m = re.match(r"^(\d+) ·", str(r[1] or ""))
    if m:
        secciones[int(m.group(1))] = i


def filas(sec):
    out, i = [], secciones[sec] + 2
    while i < len(rows):
        r = rows[i]
        if r[1] is None or not str(r[1]).strip().isdigit():
            break
        out.append(r)
        i += 1
    return out


cats, info, partes = {}, {}, {}
for sec in sorted(secciones):
    for r in filas(sec):
        cis = str(r[1]).strip()
        desc = r[5] if sec == 8 else r[4]
        cats.setdefault(cis, set()).add(sec)
        if cis not in info:
            info[cis] = {"cest": str(r[2] or ""), "desde": iso(r[3]), "desc": str(desc or "").strip()}
        if sec == 2:
            for v in (r[5], r[6]):
                p = norm_part(v) if v else None
                if p:
                    partes.setdefault(cis, []).append(p)

# 3 · Incidencias, fallas y eventos
incidencias, fallas, eventos = [], [], []
for cis in sorted(cats, key=lambda c: info[c]["desde"] or ""):
    secs, d = cats[cis], info[cis]
    u = por_cis.get(cis)
    if not u:
        u = {"id": len(unidades) + 1, "cis": cis, "check_digit": None,
             "salida_linea": d["desde"], "st": "", "ssc": "", "smo": "", "cest": d["cest"],
             "tipo": "cronos"}
        unidades.append(u)
        por_cis[cis] = u
    estado = "caso_especial" if 7 in secs else "en_espera"
    inc = {"id": len(incidencias) + 1, "unidad_id": u["id"], "estado": estado, "origen": "revision",
           "detectada_at": d["desde"], "cerrada_at": None, "notas": d["desc"]}
    incidencias.append(inc)

    def add_falla(tipo, part=None, desc=""):
        fallas.append({"id": len(fallas) + 1, "incidencia_id": inc["id"], "tipo_falla_id": tipo,
                       "particularidad_id": part, "descripcion": desc, "resuelta_at": None})

    if 6 in secs:
        add_falla(1, desc=d["desc"])
    for p in dict.fromkeys(partes.get(cis, [])):
        add_falla(2, part=p)
    if 2 in secs and cis not in partes:
        add_falla(2)
    if 4 in secs:
        add_falla(3)
    if 5 in secs:
        add_falla(4)
    if 3 in secs:
        add_falla(5)
    if "marca" in d["desc"].lower():
        add_falla(6, desc=d["desc"])
    if 7 in secs:
        add_falla(7, desc=d["desc"])
    if not any(f["incidencia_id"] == inc["id"] for f in fallas):
        add_falla(8, desc=d["desc"] or "Sin observaciones cargadas")
    eventos.append({"id": len(eventos) + 1, "incidencia_id": inc["id"], "estado_anterior": None,
                    "estado_nuevo": estado, "operario_id": None, "registrado_at": d["desde"],
                    "observacion": "Migrado desde Excel"})

colores = {}
for u in unidades:
    if u["cest"]:
        nombre, hexcolor = CEST_INFO.get(u["cest"], ("", "#A8A29A"))
        colores.setdefault(u["cest"], {"nombre": nombre, "hex": hexcolor})

seed = {
    "colores": colores, "tipos_falla": TIPOS_FALLA, "particularidades": PARTICULARIDADES,
    "operarios": OPERARIOS, "unidades": unidades, "incidencias": incidencias,
    "incidencia_fallas": fallas, "eventos": eventos,
}
destino = RAIZ / "src" / "data" / "seed.json"
destino.write_text(json.dumps(seed, ensure_ascii=False, indent=1), encoding="utf-8")
print(f"seed.json: {len(unidades)} unidades, {len(incidencias)} incidencias, "
      f"{len(fallas)} fallas, {len(eventos)} eventos -> {destino}")

if CON_SQL:
    def q(v):
        if v is None or v == "":
            return "null"
        if isinstance(v, (int, float)):
            return str(v)
        return "'" + str(v).replace("'", "''") + "'"

    ln = []
    for cest, c in sorted(colores.items()):
        ln.append(f"insert into colores (cest, nombre) values ({q(cest)}, {q(c['nombre'])}) on conflict do nothing;")
    for u in unidades:
        ln.append("insert into unidades (id, cis, check_digit, cest, st, ssc, smo, salida_linea) values "
                  f"({u['id']}, {q(u['cis'])}, {q(u['check_digit'])}, {q(u['cest'])}, {q(u['st'])}, "
                  f"{q(u['ssc'])}, {q(u['smo'])}, {q(u['salida_linea'])});")
    for i in incidencias:
        ln.append("insert into incidencias (id, unidad_id, estado, detectada_at, notas) values "
                  f"({i['id']}, {i['unidad_id']}, {q(i['estado'])}, {q(i['detectada_at'])}, {q(i['notas'])});")
    for f in fallas:
        ln.append("insert into incidencia_fallas (id, incidencia_id, tipo_falla_id, particularidad_id, descripcion) values "
                  f"({f['id']}, {f['incidencia_id']}, {f['tipo_falla_id']}, {q(f['particularidad_id'])}, {q(f['descripcion'])});")
    for e in eventos:
        ln.append("insert into eventos (incidencia_id, estado_anterior, estado_nuevo, registrado_at, observacion) values "
                  f"({e['incidencia_id']}, null, {q(e['estado_nuevo'])}, {q(e['registrado_at'])}, {q(e['observacion'])});")
    ln.append("select setval('unidades_id_seq', (select max(id) from unidades));")
    ln.append("select setval('incidencias_id_seq', (select max(id) from incidencias));")
    ln.append("select setval('incidencia_fallas_id_seq', (select max(id) from incidencia_fallas));")
    sql = RAIZ / "supabase" / "seed.sql"
    sql.write_text("\n".join(ln) + "\n", encoding="utf-8")
    print(f"seed.sql: {len(ln)} sentencias -> {sql}")
